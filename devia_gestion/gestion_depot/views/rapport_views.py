# gestion_depot/views/rapport_views.py

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, F, FloatField, Case, When, Value
from django.db.models.functions import Cast, Coalesce
from django.utils import timezone
from datetime import datetime, timedelta, time
from gestion_depot.models import BonVente, LigneVente, Produit
from gestion_depot.decorators import group_required
from collections import defaultdict
from django.utils.dateformat import DateFormat
import json
import openpyxl
from openpyxl.styles import Font, Alignment


def _calculer_rapport(date_debut, date_fin, periode):
    """Calcule l'ensemble des indicateurs du rapport pour les filtres donnés."""
    today = timezone.now().date()
    if periode == 'journalier':
        date_debut_obj = today
        date_fin_obj = today
    elif periode == 'hebdomadaire':
        date_debut_obj = today - timedelta(days=7)
        date_fin_obj = today
    elif periode == 'mensuel':
        date_debut_obj = today.replace(day=1)
        date_fin_obj = today
    else:
        date_debut_obj = date_fin_obj = None
        if date_debut:
            try:
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_fin:
            try:
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
            except ValueError:
                pass

    if date_debut_obj and date_fin_obj and date_debut_obj > date_fin_obj:
        date_debut_obj = date_fin_obj = None

    # Filtrer les bons validés
    ventes_qs = BonVente.objects.filter(statut='valide')
    if date_debut_obj:
        ventes_qs = ventes_qs.filter(date_vente__gte=date_debut_obj)
    if date_fin_obj:
        ventes_qs = ventes_qs.filter(date_vente__lte=datetime.combine(date_fin_obj, time.max))

    bon_ids = list(ventes_qs.values_list('id', flat=True))

    # Statistiques par produit (avec bénéfice)
    lignes_qs = LigneVente.objects.filter(bon_id__in=bon_ids).select_related('produit', 'bon', 'bon__vendeur')

    stats_produits = lignes_qs.values(
        'produit__id',
        'produit__nom',
        'produit__prix_achat_casier',
        'produit__prix_vente_casier'
    ).annotate(
        quantite_totale=Sum(Cast(F('quantite_casiers') * F('fraction'), FloatField())),
        nombre_ventes=Count('id'),
    ).order_by('-quantite_totale')

    stats_produits_list = []
    for stat in stats_produits:
        quantite = float(stat['quantite_totale'])
        prix_vente = float(stat['produit__prix_vente_casier'])
        prix_achat = float(stat['produit__prix_achat_casier'])

        revenu = prix_vente * quantite
        cout = prix_achat * quantite
        benefice = revenu - cout

        stats_produits_list.append({
            'produit__nom': stat['produit__nom'],
            'quantite_totale': quantite,
            'revenu_total': revenu,
            'cout_total': cout,
            'benefice_total': benefice,
            'nombre_ventes': stat['nombre_ventes'],
        })

    total_revenu = sum(s['revenu_total'] for s in stats_produits_list)
    total_cout = sum(s['cout_total'] for s in stats_produits_list)
    total_benefice = sum(s['benefice_total'] for s in stats_produits_list)

    # Produits plus/moins vendus
    produit_plus_vendu = max(stats_produits_list, key=lambda x: x['quantite_totale'], default=None)
    produit_moins_vendu = min(stats_produits_list, key=lambda x: x['quantite_totale'], default=None)

    # Inventaire avec stock annoté
    produits_stock = Produit.objects.annotate(
        stock_actuel=Coalesce(
            Sum(
                Case(
                    When(mouvement__type_mouvement='entree', then=F('mouvement__quantite_casiers')),
                    When(mouvement__type_mouvement='sortie', then=-F('mouvement__quantite_casiers')),
                    default=Value(0.0),
                    output_field=FloatField()
                )
            ),
            Value(0.0)
        )
    ).only('nom', 'seuil_alerte')

    inventaire = [
        {
            'nom': p.nom,
            'stock': round(p.stock_actuel, 2),
            'seuil_alerte': p.seuil_alerte,
            'en_alerte': p.stock_actuel <= p.seuil_alerte,
        }
        for p in produits_stock
    ]

    # Données graphique
    ventes_par_jour = defaultdict(float)
    for ligne in lignes_qs:
        jour = DateFormat(ligne.bon.date_vente).format('Y-m-d')
        total_ligne = float(ligne.produit.prix_vente_casier * ligne.quantite_casiers * ligne.fraction)
        ventes_par_jour[jour] += total_ligne

    labels_jours = sorted(ventes_par_jour.keys())
    data_ventes = [round(ventes_par_jour[jour], 0) for jour in labels_jours]

    # Lignes détaillées (avec bénéfice par ligne)
    lignes = []
    for ligne in lignes_qs.order_by('-bon__date_vente'):
        try:
            quantite_totale = float(ligne.quantite_casiers * ligne.fraction)
            ca = float(ligne.produit.prix_vente_casier) * quantite_totale
            cout = float(ligne.produit.prix_achat_casier) * quantite_totale
            benefice = ca - cout
        except (ValueError, TypeError):
            ca = 0
            cout = 0
            benefice = 0
            quantite_totale = 0

        lignes.append({
            'bon': ligne.bon,
            'produit': ligne.produit,
            'quantite_totale': quantite_totale,
            'ca': ca,
            'benefice': benefice,
            'fraction_display': str(ligne.fraction),
        })

    return {
        'total_revenu': round(total_revenu, 2),
        'total_cout': round(total_cout, 2),
        'total_benefice': round(total_benefice, 2),
        'produit_plus_vendu': produit_plus_vendu,
        'produit_moins_vendu': produit_moins_vendu,
        'inventaire': inventaire,
        'date_debut': date_debut_obj,
        'date_fin': date_fin_obj,
        'periode': periode,
        'labels_jours': json.dumps(labels_jours),
        'data_ventes': json.dumps(data_ventes),
        'lignes': lignes,
        'stats_produits': stats_produits_list,
    }


def _produit_stat_json(stat):
    """Normalise un dict produit (clés produit__nom) pour le JSON."""
    if stat is None:
        return None
    return {
        'nom': stat['produit__nom'],
        'quantite_totale': round(stat['quantite_totale'], 2),
        'revenu_total': round(stat['revenu_total'], 0),
        'cout_total': round(stat['cout_total'], 0),
        'benefice_total': round(stat['benefice_total'], 0),
        'nombre_ventes': stat['nombre_ventes'],
    }


@login_required
@group_required('Gérant', 'Admin')
def rapport_ventes(request):
    context = _calculer_rapport(
        request.GET.get('date_debut'),
        request.GET.get('date_fin'),
        request.GET.get('periode'),
    )

    # Export Excel
    if request.GET.get('export') == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="rapport_ventes.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport de Ventes"

        ws.merge_cells('A1:F1')
        ws['A1'] = "RAPPORT DE VENTES"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        debut_str = str(context['date_debut']) if context['date_debut'] else "Indéfini"
        fin_str = str(context['date_fin']) if context['date_fin'] else "Indéfini"
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Période : du {debut_str} au {fin_str}"
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.append([])
        ws.append(['Produit', 'Quantité', 'CA (FCFA)', 'Coût (FCFA)', 'Bénéfice (FCFA)', 'Ventes'])
        for cell in ws[4]:
            cell.font = Font(bold=True)

        for stat in context['stats_produits']:
            ws.append([
                stat['produit__nom'],
                round(stat['quantite_totale'], 2),
                round(stat['revenu_total'], 0),
                round(stat['cout_total'], 0),
                round(stat['benefice_total'], 0),
                stat['nombre_ventes']
            ])

        ws.append([])
        ws.append(['TOTAL', '',
                   round(context['total_revenu'], 0),
                   round(context['total_cout'], 0),
                   round(context['total_benefice'], 0),
                   ''])

        wb.save(response)
        return response

    return render(request, 'gestion_depot/rapport.html', context)


@login_required
@group_required('Gérant', 'Admin')
def rapport_ventes_ajax(request):
    """Renvoie les indicateurs du rapport en JSON pour des filtres dynamiques."""
    context = _calculer_rapport(
        request.GET.get('date_debut'),
        request.GET.get('date_fin'),
        request.GET.get('periode'),
    )

    payload = {
        'total_revenu': context['total_revenu'],
        'total_cout': context['total_cout'],
        'total_benefice': context['total_benefice'],
        'produit_plus_vendu': _produit_stat_json(context['produit_plus_vendu']),
        'produit_moins_vendu': _produit_stat_json(context['produit_moins_vendu']),
        'stats_produits': [_produit_stat_json(s) for s in context['stats_produits']],
        'lignes': [
            {
                'date': DateFormat(l['bon'].date_vente).format('d/m/Y'),
                'produit': l['produit'].nom,
                'quantite_totale': round(l['quantite_totale'], 2),
                'ca': round(l['ca'], 0),
                'benefice': round(l['benefice'], 0),
                'vendeur': l['bon'].vendeur.username if l['bon'].vendeur else '-',
            }
            for l in context['lignes']
        ],
        'inventaire': context['inventaire'],
        'labels_jours': json.loads(context['labels_jours']),
        'data_ventes': json.loads(context['data_ventes']),
    }
    return JsonResponse(payload)
